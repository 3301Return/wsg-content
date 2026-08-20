#!/usr/bin/env python3
"""
Content Calendar Parser
=======================

Reads the WSG keyword/calendar xlsx and outputs a clean Markdown file the
article generator can read on demand.

Usage:
    python3 scripts/parse_calendar.py

Output:
    content-calendar/calendar.md  (next-up + full pending list)
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Need: pip install openpyxl --break-system-packages")
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_GLOB = "content-calendar/*Keyword*Prioritization*.xlsx"


def find_xlsx():
    matches = list(PROJECT_ROOT.glob(XLSX_GLOB))
    if not matches:
        print(f"Could not find calendar xlsx at {XLSX_GLOB}")
        sys.exit(1)
    return matches[0]


def cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return str(v).strip() if v is not None else None


def parse_writers_sheet(ws):
    """Parse the 'writers and topics' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = rows[0]
    # Find column indexes
    def col(name):
        for i, h in enumerate(header):
            if h and name.lower() in str(h).lower():
                return i
        return None
    c_title = col("Blog Post Title")
    c_keyword = col("Keyword")
    c_writer = col("Best Writer")
    c_reason = col("Reason")
    c_industry = col("Writer industry")
    c_status = col("week 1 status")

    sent = []
    pending = []
    seen_for_later = False
    for row in rows[1:]:
        title = cell(row, c_title)
        if not title:
            continue
        if title.upper().strip() == "FOR LATER":
            seen_for_later = True
            continue
        record = {
            "title": title,
            "keyword": cell(row, c_keyword),
            "writer": cell(row, c_writer),
            "reason": cell(row, c_reason),
            "industry": cell(row, c_industry),
            "status": cell(row, c_status),
        }
        if record["status"] and record["status"].lower() == "sent":
            sent.append(record)
        else:
            pending.append(record)
    return sent, pending


def parse_master_sheet(ws):
    """Parse the main keyword sheet (priority, search volume)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    def col(name):
        for i, h in enumerate(header):
            if h and name.lower() in str(h).lower():
                return i
        return None
    c_keyword_parent = col("Main/parent Keyword")
    c_keyword_cluster = col("Cluster Keyword")
    c_us = col("US Traffic")
    c_difficulty = col("KW Difficulty")
    c_priority = col("Prioritized")
    c_post = col("Potential blog post")
    out = []
    for row in rows[1:]:
        keyword = cell(row, c_keyword_parent) or cell(row, c_keyword_cluster)
        if not keyword:
            continue
        rec = {
            "keyword": keyword,
            "us_traffic": cell(row, c_us),
            "difficulty": cell(row, c_difficulty),
            "prioritized": cell(row, c_priority),
            "post_idea": cell(row, c_post),
        }
        if rec["prioritized"] and rec["prioritized"].lower() == "yes":
            out.append(rec)
    return out


def format_record(rec):
    parts = [f"### {rec['title']}", ""]
    if rec.get("keyword"):
        parts.append(f"- **Keyword:** {rec['keyword']}")
    if rec.get("writer"):
        parts.append(f"- **Writer:** {rec['writer']}")
    if rec.get("industry"):
        parts.append(f"- **Writer industry:** {rec['industry']}")
    if rec.get("reason"):
        parts.append(f"- **Why this writer:** {rec['reason']}")
    parts.append("")
    return "\n".join(parts)


def main():
    xlsx = find_xlsx()
    print(f"Reading {xlsx}...")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    sent, pending = [], []
    for sheet_name in wb.sheetnames:
        if "writers and topics" in sheet_name.lower():
            sent, pending = parse_writers_sheet(wb[sheet_name])
            break

    prioritized = []
    for sheet_name in wb.sheetnames:
        if "master" in sheet_name.lower() or "prioritization" in sheet_name.lower():
            prioritized = parse_master_sheet(wb[sheet_name])
            break

    out = []
    out.append("# Content Calendar")
    out.append("")
    out.append(
        "_Auto-generated from the Keyword Prioritization xlsx by `scripts/parse_calendar.py`. "
        "Do not edit by hand — edit the xlsx and re-run the script._"
    )
    out.append("")
    out.append("## Next up (pending, not yet sent)")
    out.append("")
    if not pending:
        out.append("_No pending articles._")
    else:
        for rec in pending:
            out.append(format_record(rec))
    out.append("")
    out.append(f"## Already sent ({len(sent)} articles)")
    out.append("")
    for rec in sent:
        out.append(f"- {rec['title']} — {rec.get('writer', 'unassigned')}")
    out.append("")
    out.append("## High-priority keywords without an assigned article (top 20)")
    out.append("")
    out.append("| Keyword | US Traffic | Difficulty | Post idea |")
    out.append("|---|---|---|---|")
    for rec in prioritized[:20]:
        out.append(
            f"| {rec['keyword']} | {rec['us_traffic'] or '-'} | "
            f"{rec['difficulty'] or '-'} | {rec['post_idea'] or '-'} |"
        )
    out.append("")

    out_path = PROJECT_ROOT / "content-calendar" / "calendar.md"
    out_path.write_text("\n".join(out), encoding="utf-8")
    print(f"Wrote {out_path}")
    print(f"  - {len(pending)} pending articles")
    print(f"  - {len(sent)} already sent")
    print(f"  - {len(prioritized)} prioritized keywords parsed")


if __name__ == "__main__":
    main()
